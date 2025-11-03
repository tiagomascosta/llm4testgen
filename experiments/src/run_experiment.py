import os
import json
import subprocess
import time
import shutil
import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime

CONFIG_PATH = Path(__file__).parent / "sample_config_batch2.json"
OUTPUT_ROOT = Path(__file__).parent / "output"
RESULTS_ROOT = Path(__file__).parent / "results"
TOOL_PATH = Path(__file__).parent.parent.parent / "generate_test_suite.py"
TEMPLATE_PATH = Path(__file__).parent / "experiment_sheet_template.xlsx"

# Configuration: Set to None to start from beginning, or specify repo name to start from
START_FROM_REPO = "jaxb-xew-plugin" # Test with easiest repository

# Number of times to run each repository
RUNS_PER_REPO = 5  # Just one run for testing

# Tool configuration parameters
MAX_FIX_ATTEMPTS = 8
MAX_RUNTIME_FIX_ATTEMPTS = 8
PORT = 11435  # Default port for Ollama

# Model configuration
CODE_MODEL = "qwen3-coder:30b"
NON_CODE_MODEL = "qwen3-coder:30b"
NON_CODE_MODEL_BUG = "qwen3-coder:30b"

# Fix loop configuration
MAX_RUNTIME_FIX_EXAMPLES = 3
MAX_COMPILE_FIX_EXAMPLES = 3
MAX_SCAFFOLD_EXAMPLES = 3

# Directory configuration
INPUT_DIR = "/home/tiago/Desktop/Faculdade/Thesis/implementation/input4"  # Set to custom path for repository downloads, e.g., "/tmp/experiment_repos"
OUTPUT_DIR = "/home/tiago/Desktop/Faculdade/Thesis/implementation/output4"  # Set to custom path for outputs, e.g., "/tmp/experiment_outputs"

# Force specific experiment number (set to None for auto-increment)
FORCE_EXPERIMENT_NUMBER = 10

# Sample data for prefilling
SAMPLE_DATA_PATH = Path(__file__).parent.parent / "dataset" / "src" / "sample.csv" 


def load_config():
    with open(CONFIG_PATH) as f:
        config = json.load(f)
    for entry in config:
        if not all(k in entry for k in ("repo_url", "commit_hash", "method")):
            raise ValueError(f"Missing required fields in config entry: {entry}")
    return config


def load_sample_data():
    """Load sample data for prefilling Excel sheet."""
    df = pd.read_csv(SAMPLE_DATA_PATH)
    # Create a mapping from repo_url to sample data
    sample_data = {}
    for _, row in df.iterrows():
        sample_data[row['repo_url']] = row.to_dict()
    return sample_data


def get_sample_data_for_repo(sample_data, repo_url):
    """Get sample data for a specific repository."""
    return sample_data.get(repo_url, {})


def get_next_experiment_number():
    """Get the next available experiment number by checking existing files."""
    if FORCE_EXPERIMENT_NUMBER is not None:
        return FORCE_EXPERIMENT_NUMBER
    
    experiment_number = 1
    while (RESULTS_ROOT / f"experiment_{experiment_number}.xlsx").exists():
        experiment_number += 1
    return experiment_number


def find_first_empty_row(experiment_file):
    """Find the first empty row in the Excel file."""
    if not experiment_file.exists():
        return 3  # Start from row 3 if file doesn't exist
    
    wb = openpyxl.load_workbook(experiment_file)
    ws = wb['Results']
    
    # Start from row 3 (after headers)
    row = 3
    while ws.cell(row=row, column=1).value is not None:  # Check if first column has data
        row += 1
    
    wb.close()
    return row


def create_experiment_sheet():
    """Create a copy of the template for this experiment."""
    experiment_number = get_next_experiment_number()
    experiment_file = RESULTS_ROOT / f"experiment_{experiment_number}.xlsx"
    
    # Only copy template if file doesn't exist (for forced experiment numbers)
    if not experiment_file.exists():
        shutil.copy2(TEMPLATE_PATH, experiment_file)
        print(f"Created experiment sheet: {experiment_file}")
    else:
        print(f"Using existing experiment sheet: {experiment_file}")
    
    return experiment_number, experiment_file


def prefilled_excel_row(experiment_file, row_number, repo_data, sample_data, run_num, repo_index):
    """Prefill a row in the Excel sheet with metadata before running the experiment."""
    wb = openpyxl.load_workbook(experiment_file)
    ws = wb['Results']
    
    # Get sample data for this repository
    repo_url = repo_data["repo_url"]
    sample_row = get_sample_data_for_repo(sample_data, repo_url)
    
    # Prefill metadata (columns A-Q)
    # A: Run number
    ws.cell(row=row_number, column=1, value=run_num)
    
    # B: Project name
    ws.cell(row=row_number, column=2, value=sample_row.get('project', ''))
    
    # C: Repo URL
    ws.cell(row=row_number, column=3, value=repo_url)
    
    # D: Bug file path
    ws.cell(row=row_number, column=4, value=sample_row.get('bug_file_path', ''))
    
    # E: Bug commit hash
    ws.cell(row=row_number, column=5, value=repo_data["commit_hash"])
    
    # F: Fix commit hash
    ws.cell(row=row_number, column=6, value=repo_data["fix_commit_hash"])
    
    # G: Method name
    ws.cell(row=row_number, column=7, value=sample_row.get('method_name', ''))
    
    # H: Repo Java Version (will be filled after run)
    # I: Build System (will be filled after run)
    
    # N: LOC
    ws.cell(row=row_number, column=14, value=sample_row.get('loc', ''))
    
    # O: Cyclomatic Complexity
    ws.cell(row=row_number, column=15, value=sample_row.get('cyclomatic_complexity', ''))
    
    # P: Halstead Volume
    ws.cell(row=row_number, column=16, value=sample_row.get('halstead_volume', ''))
    
    # Q: Maintainability Index
    ws.cell(row=row_number, column=17, value=sample_row.get('maintainability_index', ''))
    
    # Save the workbook
    wb.save(experiment_file)
    wb.close()
    
    print(f"    Prefilled row {row_number} with metadata")


def parse_tool_report(repo_name, repo_output_dir):
    """Parse the JSON report from the tool output."""
    report_path = repo_output_dir / "reports" / f"{repo_name}_report.json"
    
    if not report_path.exists():
        print(f"    Warning: Report file not found: {report_path}")
        return None
    
    try:
        with open(report_path, 'r') as f:
            report = json.load(f)
        return report
    except Exception as e:
        print(f"    Error reading report: {e}")
        return None


def fill_post_run_data(experiment_file, row_number, repo_name, repo_output_dir, elapsed_time):
    """Fill the Excel row with post-run data from the tool report."""
    wb = openpyxl.load_workbook(experiment_file)
    ws = wb['Results']
    
    # Parse the tool report
    report = parse_tool_report(repo_name, repo_output_dir)
    
    if report is None:
        print(f"    Warning: Could not parse report for {repo_name}")
        wb.save(experiment_file)
        wb.close()
        return
    
    # Extract data from the report
    try:
        # Tool Configuration (columns H-M)
        # H: Repo Java Version
        java_version = report.get('java_repo_version', '')
        ws.cell(row=row_number, column=8, value=java_version)
        
        # I: Build System
        build_system = report.get('build_system', '')
        ws.cell(row=row_number, column=9, value=build_system)
        
        # J: Non-Code LLM
        non_code_llm = report.get('llm_models', {}).get('non_code_tasks', '')
        ws.cell(row=row_number, column=10, value=non_code_llm)
        
        # K: Code LLM
        code_llm = report.get('llm_models', {}).get('code_tasks', '')
        ws.cell(row=row_number, column=11, value=code_llm)
        
        # L: # Fix Iterations - from CLI option max_fix_attempts
        fix_iterations = report.get('cli_options', {}).get('max_fix_attempts', 0)
        ws.cell(row=row_number, column=12, value=fix_iterations)
        
        # M: # Fix Examples - from CLI option max_compile_fix_examples
        fix_examples = report.get('cli_options', {}).get('max_compile_fix_examples', 0)
        ws.cell(row=row_number, column=13, value=fix_examples)
        
        # Test Scenarios (columns R-T)
        # R: # Scenarios
        scenarios = report.get('test_scenarios', {}).get('raw_scenarios', 0)
        ws.cell(row=row_number, column=18, value=scenarios)
        
        # S: # Clustered Scenarios
        clustered_scenarios = report.get('test_scenarios', {}).get('total_clustered', 0)
        ws.cell(row=row_number, column=19, value=clustered_scenarios)
        
        # T: Total # Fix Attempts - sum of all fix attempts for all test cases
        test_generation = report.get('test_generation', {})
        scenarios_data = test_generation.get('scenarios', {})
        total_fix_attempts = sum(
            scenario.get('fix_attempts', 0) 
            for scenario in scenarios_data.values()
        )
        ws.cell(row=row_number, column=20, value=total_fix_attempts)
        
        # Test Generation (columns U-Y)
        # U: # Compiled Tests - count scenarios where compiled is true
        test_generation = report.get('test_generation', {})
        scenarios_data = test_generation.get('scenarios', {})
        compiled_tests = sum(1 for scenario in scenarios_data.values() if scenario.get('compiled', False))
        ws.cell(row=row_number, column=21, value=compiled_tests)
        
        # V: # Non-Compiled Tests - count scenarios where compiled is false
        non_compiled_tests = sum(1 for scenario in scenarios_data.values() if not scenario.get('compiled', False))
        ws.cell(row=row_number, column=22, value=non_compiled_tests)
        
        # W: # Assertions
        assertions = report.get('final_test_suite', {}).get('assertions', 0) or 0
        ws.cell(row=row_number, column=23, value=assertions)
        
        # X: # Test Cases
        test_cases = report.get('final_test_suite', {}).get('tests_in_final_test_suite', 0) or 0
        ws.cell(row=row_number, column=24, value=test_cases)
        
        # Y: Names of test cases
        test_names = report.get('final_test_suite', {}).get('final_test_names', [])
        test_names_str = ', '.join(test_names) if test_names else ''
        ws.cell(row=row_number, column=25, value=test_names_str)
        
        # Test Effectiveness (columns Z-AD)
        # Z: Bug Detected?
        bug_detected = report.get('bug_assessment', {}).get('bug_revealed', False)
        ws.cell(row=row_number, column=26, value=bug_detected)
        
        # AA: Regression?
        regression = report.get('regression_detection', {}).get('regression_detected', False)
        ws.cell(row=row_number, column=27, value=regression)
        
        # Additional Bug Assessment Fields
        # AW: Number of tests that reveal bugs (count of bug_revealing_test_names)
        bug_revealing_test_names = report.get('bug_assessment', {}).get('bug_revealing_test_names', [])
        bug_revealing_tests_count = len(bug_revealing_test_names) if bug_revealing_test_names else 0
        ws.cell(row=row_number, column=49, value=bug_revealing_tests_count)
        
        # AX: Number of potential bug revealing tests
        potential_bug_revealing_tests = report.get('bug_assessment', {}).get('potential_bug_revealing_tests', 0) or 0
        ws.cell(row=row_number, column=50, value=potential_bug_revealing_tests)
        
        # AB: Line Cov. (%)
        coverage_data = report.get('coverage', {})
        lines_covered = coverage_data.get('lines_covered', 0) or 0
        lines_total = coverage_data.get('lines_total', 0) or 0
        line_coverage = round((lines_covered / lines_total) * 100, 2) if lines_total > 0 else 0
        ws.cell(row=row_number, column=28, value=line_coverage)
        
        # AC: Branch Cov. (%)
        branches_covered = coverage_data.get('branches_covered', 0) or 0
        branches_total = coverage_data.get('branches_total', 0) or 0
        if branches_total == 0 or branches_total is None:
            # Only consider 100% coverage if there are test cases AND no branches exist
            if test_cases > 0 and branches_total is not None:
                branch_coverage = 100.0  # 100% coverage when no branches exist but tests were run
            else:
                branch_coverage = 0.0  # 0% coverage when no test cases were generated
        else:
            branch_coverage = round((branches_covered / branches_total) * 100, 2)
        ws.cell(row=row_number, column=29, value=branch_coverage)
        
        # AD: Instruction Cov. (%)
        instructions_covered = coverage_data.get('instructions_covered', 0) or 0
        instructions_total = coverage_data.get('instructions_total', 0) or 0
        if instructions_total == 0:
            instruction_coverage = 0
        else:
            instruction_coverage = round((instructions_covered / instructions_total) * 100, 2)
        ws.cell(row=row_number, column=30, value=instruction_coverage)
        
        # Individual Test Execution (columns AE-AK)
        # AE: # Assertion Errors
        assertion_errors = report.get('test_execution', {}).get('individual', {}).get('assertion_errors', 0)
        ws.cell(row=row_number, column=31, value=assertion_errors)
        
        # AF: # Runtime Errors
        runtime_errors = report.get('test_execution', {}).get('individual', {}).get('runtime_errors', 0)
        ws.cell(row=row_number, column=32, value=runtime_errors)
        
        # AG: # Bug Revealing RE
        bug_revealing_runtime_errors = report.get('test_execution', {}).get('individual', {}).get('bug_revealing_runtime_errors', 0)
        ws.cell(row=row_number, column=33, value=bug_revealing_runtime_errors)
        
        # AH: # Fixable RE
        fixable_runtime_errors = report.get('test_execution', {}).get('individual', {}).get('fixable_runtime_errors', 0)
        ws.cell(row=row_number, column=34, value=fixable_runtime_errors)
        
        # AI: # Timeout
        timeout_errors = report.get('test_execution', {}).get('individual', {}).get('timeout_errors', 0)
        ws.cell(row=row_number, column=35, value=timeout_errors)
        
        # AJ: # Number of RFL Attempts
        total_rfl_attempts = report.get('test_execution', {}).get('individual', {}).get('total_rfl_attempts', 0)
        ws.cell(row=row_number, column=36, value=total_rfl_attempts)
        
        # AK: # Tests Fixed
        total_tests_fixed = report.get('test_execution', {}).get('individual', {}).get('total_tests_fixed', 0)
        ws.cell(row=row_number, column=37, value=total_tests_fixed)
        
        # Summary Test Execution (columns AL-AN)
        # AL: # Assertion Errors
        summary_assertion_errors = report.get('test_execution', {}).get('summary', {}).get('assertion_errors', 0)
        ws.cell(row=row_number, column=38, value=summary_assertion_errors)
        
        # AM: # Runtime Errors
        summary_runtime_errors = report.get('test_execution', {}).get('summary', {}).get('runtime_errors', 0)
        ws.cell(row=row_number, column=39, value=summary_runtime_errors)
        
        # AN: # Timeout errors
        summary_timeout_errors = report.get('test_execution', {}).get('summary', {}).get('timeout_errors', 0)
        ws.cell(row=row_number, column=40, value=summary_timeout_errors)
        
        # Bug Hunting Test Generation (columns AO-AR)
        # AO: # Scenarios
        bug_hunting_scenarios = report.get('bug_hunting_test_generation', {}).get('total_scenarios', 0)
        ws.cell(row=row_number, column=41, value=bug_hunting_scenarios)
        
        # AP: Total # Fix Attempts
        total_fix_attempts = sum(
            scenario.get('fix_attempts', 0) or 0
            for scenario in report.get('bug_hunting_test_generation', {}).get('scenarios', {}).values()
        )
        ws.cell(row=row_number, column=42, value=total_fix_attempts)
        
        # AQ: # Compiled Tests
        compiled_tests = sum(1 for scenario in report.get('bug_hunting_test_generation', {}).get('scenarios', {}).values() if scenario.get('compiled', False))
        ws.cell(row=row_number, column=43, value=compiled_tests)
        
        # AR: # Non-Compiled Tests
        non_compiled_tests = sum(1 for scenario in report.get('bug_hunting_test_generation', {}).get('scenarios', {}).values() if not scenario.get('compiled', False))
        ws.cell(row=row_number, column=44, value=non_compiled_tests)
        
        # Time (columns AS-AU)
        # AS: Elapsed Time
        elapsed_time_from_json = report.get('elapsed_time', 0)
        ws.cell(row=row_number, column=45, value=elapsed_time_from_json)
        
        # AT: LLM Response Time
        llm_response_time = report.get('llm_response_time', 0)
        ws.cell(row=row_number, column=46, value=llm_response_time)
        
        # AU: # LLM Requests
        llm_requests = report.get('llm_requests', 0)
        ws.cell(row=row_number, column=47, value=llm_requests)
        
        # Others (column AV)
        # AV: # First Try Compilation (regular + bug hunting scenarios)
        # Count regular scenarios
        test_generation = report.get('test_generation', {})
        regular_scenarios = test_generation.get('scenarios', {})
        regular_first_try = sum(1 for scenario in regular_scenarios.values() if scenario.get('compiled_on_first_attempt', False))
        
        # Count bug hunting scenarios
        bug_hunting_test_generation = report.get('bug_hunting_test_generation', {})
        bug_hunting_scenarios = bug_hunting_test_generation.get('scenarios', {})
        bug_hunting_first_try = sum(1 for scenario in bug_hunting_scenarios.values() if scenario.get('compiled_on_first_attempt', False))
        
        # Total first try compilation (both regular and bug hunting)
        first_try_compilation = regular_first_try + bug_hunting_first_try
        ws.cell(row=row_number, column=48, value=first_try_compilation)
        
        
        print(f"    Filled row {row_number} with post-run data")
        
    except Exception as e:
        print(f"    Error filling post-run data: {e}")
    
    # Save the workbook
    wb.save(experiment_file)
    wb.close()


def save_experiment_sheet(experiment_file):
    """Save the experiment sheet (called after each run)."""
    # The workbook is already saved in prefilled_excel_row, but this function
    # can be used for additional saves if needed
    pass


def run_for_repo(repo_cfg):
    repo_url = repo_cfg["repo_url"]
    commit_hash = repo_cfg["commit_hash"]
    fix_commit_hash = repo_cfg["fix_commit_hash"]
    method = repo_cfg["method"]
    repo_name = repo_url.replace(".git", "").split("/")[-1]
    # Use custom OUTPUT_DIR if specified, otherwise use OUTPUT_ROOT
    if OUTPUT_DIR:
        output_dir = Path(OUTPUT_DIR).resolve()
        repo_output_dir = Path(OUTPUT_DIR) / repo_name
    else:
        output_dir = OUTPUT_ROOT.resolve()  # Use OUTPUT_ROOT directly, not a subdirectory
        repo_output_dir = OUTPUT_ROOT / repo_name  # Where the tool should create its output
    log_path = repo_output_dir / "run_output.log"  # Log goes in the repo folder created by the tool

    # Change to the implementation directory before running the tool
    implementation_dir = TOOL_PATH.parent
    cmd = [
        "python", "-u", str(TOOL_PATH.name),  # Use just the filename, not full path
        "--repo-url", repo_url,
        "--commit-hash", commit_hash,
        "--fix-commit-hash", fix_commit_hash,
        "--method", method,
        "--output-dir", str(output_dir),
        "--max-fix-attempts", str(MAX_FIX_ATTEMPTS),
        "--max-runtime-fix-attempts", str(MAX_RUNTIME_FIX_ATTEMPTS),
        "--ollama-port", str(PORT),
        "--code-model", CODE_MODEL,
        "--non-code-model", NON_CODE_MODEL,
        "--non-code-model-bug", NON_CODE_MODEL_BUG,
        "--max-runtime-fix-examples", str(MAX_RUNTIME_FIX_EXAMPLES),
        "--max-compile-fix-examples", str(MAX_COMPILE_FIX_EXAMPLES),
        "--max-scaffold-examples", str(MAX_SCAFFOLD_EXAMPLES)
    ]
    
    # Add input directory if specified
    if INPUT_DIR:
        cmd.extend(["--input-dir", str(INPUT_DIR)])

    start_time = time.time()
    
    # Run the tool
    try:
        result = subprocess.run(cmd, text=True, timeout=60*60*24, cwd=implementation_dir)
        exit_code = result.returncode
    except Exception as e:
        result = type('obj', (object,), {'stdout': '', 'stderr': f'Exception: {e}', 'returncode': -1})()
        exit_code = -1
    
    elapsed = time.time() - start_time
    
    # Ensure the repo output directory exists (in case tool didn't create it)
    repo_output_dir.mkdir(parents=True, exist_ok=True)
    
    # Write the captured output to the log file
    with open(log_path, "w") as log_file:
        log_file.write(f"Started at: {datetime.now()}\n")
        log_file.write(f"Command: {' '.join(cmd)}\n")
        if hasattr(result, 'stdout') and result.stdout:
            log_file.write(result.stdout)
        if hasattr(result, 'stderr') and result.stderr:
            log_file.write(result.stderr)
        log_file.write(f"\nExit code: {exit_code}\n")
        log_file.write(f"Elapsed time: {elapsed:.2f} seconds\n")
    
    return {
        "repo_name": repo_name,
        "exit_code": exit_code,
        "elapsed": elapsed,
        "log_path": str(log_path),
        "repo_output_dir": str(repo_output_dir)
    }


def main():
    config = load_config()
    sample_data = load_sample_data()
    OUTPUT_ROOT.mkdir(exist_ok=True)
    RESULTS_ROOT.mkdir(exist_ok=True)
    
    # Create experiment sheet
    experiment_number, experiment_file = create_experiment_sheet()
    
    summary = []
    failed = []
    total_start = time.time()
    total_repos = len(config)
    
    # Find starting index if START_FROM_REPO is specified
    start_index = 0
    if START_FROM_REPO:
        for i, repo_cfg in enumerate(config):
            repo_name = repo_cfg["repo_url"].replace(".git", "").split("/")[-1]
            if repo_name == START_FROM_REPO:
                start_index = i
                print(f"Starting from repository: {START_FROM_REPO} (index {i})")
                break
        else:
            print(f"Warning: Repository '{START_FROM_REPO}' not found in config. Starting from beginning.")
            start_index = 0
    
    # Calculate actual remaining repositories and runs
    remaining_repos = total_repos - start_index
    total_runs = remaining_repos * RUNS_PER_REPO
    
    print(f"Starting experiment {experiment_number}")
    print(f"Running {remaining_repos} repositories, {RUNS_PER_REPO} times each = {total_runs} total runs")
    print("=" * 60)
    
    run_counter = 0
    
    # Determine starting Excel row
    if FORCE_EXPERIMENT_NUMBER is not None:
        excel_row = find_first_empty_row(experiment_file)
        print(f"Starting from Excel row {excel_row} (first empty row)")
    else:
        excel_row = 3  # Start from row 3 (after headers)
    
    for idx, repo_cfg in enumerate(config[start_index:], start_index + 1):
        repo_url = repo_cfg["repo_url"]
        repo_name = repo_url.replace(".git", "").split("/")[-1]
        
        print(f"\n[{idx}/{total_repos}] Repository: {repo_name}")
        print("-" * 40)
        
        repo_results = []
        successful_runs = 0
        run_num = 1
        
        while successful_runs < RUNS_PER_REPO:
            run_counter += 1
            print(f"  Run {run_num}/{RUNS_PER_REPO} (Overall: {run_counter}/{total_runs})")
            
            # Don't prefill Excel row yet - wait to see if run succeeds
            
            start = time.time()
            result = run_for_repo(repo_cfg)
            result["run_number"] = run_num
            result["repo_index"] = idx
            result["attempt_number"] = 1  # Track which attempt this is

            elapsed = time.time() - start
            
            print(f"  Run {run_num} completed in {elapsed:.1f}s (exit code: {result['exit_code']})")
            
            # If successful, add to results and populate Excel
            if result["exit_code"] == 0:
                # Prefill Excel row with metadata for successful run
                prefilled_excel_row(experiment_file, excel_row, repo_cfg, sample_data, run_num, idx)
                
                # Fill post-run data in Excel for successful run
                fill_post_run_data(experiment_file, excel_row, repo_name, Path(result["repo_output_dir"]), elapsed)
                repo_results.append(result)
                successful_runs += 1
                run_num += 1
                excel_row += 1  # Move to next Excel row
            else:
                # Failed run - retry up to 3 times
                retry_count = 0
                max_retries = 3
                
                while result["exit_code"] != 0 and retry_count < max_retries:
                    retry_count += 1
                    print(f"    Retry {retry_count}/{max_retries} for run {run_num}")
                    
                    start = time.time()
                    result = run_for_repo(repo_cfg)
                    result["run_number"] = run_num
                    result["attempt_number"] = retry_count + 1

                    elapsed = time.time() - start
                    
                    print(f"    Retry {retry_count} completed in {elapsed:.1f}s (exit code: {result['exit_code']})")
                
                # After retries, check if we succeeded
                if result["exit_code"] == 0:
                    # Success after retry - populate Excel
                    prefilled_excel_row(experiment_file, excel_row, repo_cfg, sample_data, run_num, idx)
                    fill_post_run_data(experiment_file, excel_row, repo_name, Path(result["repo_output_dir"]), elapsed)
                    repo_results.append(result)
                    successful_runs += 1
                    run_num += 1
                    excel_row += 1  # Move to next Excel row
                else:
                    # Still failed after retries - don't populate Excel
                    failed.append(f"{repo_name}_run_{run_num}")
                    run_num += 1
                    # Don't increment excel_row - failed runs don't get Excel rows
        
        # Add repository summary
        successful_runs_count = sum(1 for r in repo_results if r["exit_code"] == 0)
        failed_runs_count = sum(1 for r in repo_results if r["exit_code"] != 0)
        total_attempts = sum(r.get("attempt_number", 1) for r in repo_results)
        
        repo_summary = {
            "repo_name": repo_name,
            "repo_index": idx,
            "runs": repo_results,
            "successful_runs": successful_runs_count,
            "failed_runs": failed_runs_count,
            "total_attempts": total_attempts,
            "avg_elapsed": sum(r["elapsed"] for r in repo_results) / len(repo_results)
        }
        summary.append(repo_summary)
    
    total_elapsed = time.time() - total_start

    # Write summary file (JSON only)
    summary_data = {
        "experiment_number": experiment_number,
        "experiment_file": str(experiment_file),
        "config": {
            "total_repos": total_repos,
            "runs_per_repo": RUNS_PER_REPO,
            "total_runs": total_runs
        },
        "results": summary,
        "total_elapsed": total_elapsed,
        "failed_runs": failed
    }
    
    with open(RESULTS_ROOT / f"experiment_{experiment_number}_summary.json", "w") as f:
        json.dump(summary_data, f, indent=2)

    print(f"\n" + "=" * 60)
    print(f"Experiment {experiment_number} complete!")
    print(f"Total repositories: {total_repos}")
    print(f"Total runs: {total_runs}")
    print(f"Failed runs: {len(failed)}")
    print(f"Total time: {total_elapsed:.1f}s")
    print(f"Results saved to: {RESULTS_ROOT / f'experiment_{experiment_number}_summary.json'}")
    print(f"Experiment sheet: {experiment_file}")


if __name__ == "__main__":
    main() 